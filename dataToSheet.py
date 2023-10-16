#!/usr/bin/env python
'''dataToSheet.py: formatted data pushed to remote sheet'''
'''WARNING: THIS USES AN OPEN SOURCE LIBRARY VULNERABLE TO BILLION LAUGHS'''

__author__ = 'ISAIAH COLEMAN'
__credits__ = ['Isaiah Coleman']
__status__ = 'Prototype'

#imports
from datetime import datetime
from openpyxl.styles import Font
from openpyxl import load_workbook
from openpyxl import Workbook
import random

#table bounds
MAX_ROW = 26
MAX_COL = 20
MIN_VAL = 1
MAX_VAL = 10

#set active workbook(can be new or pre-existing)
#wb = Workbook()
wb = load_workbook('test.xlsx')

#assigns active worksheet based on last used
ws = wb.active

#renames the active ws
ws.title = 'rainbow'

row = [None]*MAX_ROW

x, y = (0, 0)
for x in range(MAX_ROW):
    for y in range(MAX_COL):
        row[y] = random.randint(MIN_VAL, MAX_VAL)
    print(row)
    ws.append(row)

#setup font for ROW 1
font = Font(bold=True)
for row in ws['A1:T1']:
    for cell in row:
        cell.font = font

wb.save('test.xlsx')